from itertools import product
import gurobipy as gp
import pandas as pd
from openpyxl import load_workbook
import os
import random
from datetime import datetime
from itertools import product

current_time = datetime.now()
timestamp = current_time.timestamp()
random.seed(int(timestamp))

# 0. 读⼊⽂件
data_1 = pd.read_excel(r"C:\Users\86181\Desktop\数模作业、笔记\农作物种植策略\附件1修改版.xlsx", sheet_name="乡村的现有耕地")
data_2 = pd.read_excel(r"C:\Users\86181\Desktop\数模作业、笔记\农作物种植策略\附件2修改版.xlsx", sheet_name="2023年的农作物种植情况")
data_2 = data_2.ffill()# 处理合并单元格的问题
data_3 = pd.read_excel(r"C:\Users\86181\Desktop\数模作业、笔记\农作物种植策略\附件2修改版.xlsx", sheet_name="2023年统计的相关数据")
trans_dict = {"A": "平旱地", "B": "梯田", "C": "山坡地", "D": "水浇地", "E": "普通大棚", "F": "智慧大棚"}#创建字典键值对

import pandas as pd

# 假设 data_3 是一个 DataFrame
pd.set_option('display.max_rows', None)  # 显示所有行
pd.set_option('display.max_columns', None)  # 显示所有列
pd.set_option('display.width', None)  # 自动调整宽度

# 1. 输⼊参数
# 1.1 销量
q_0 = {(data_2["种植地块"][i], data_2["作物编号"][i],
        2 if data_2["种植季次"][i] == "第二季" else 1):
            data_2['种植面积/亩'][i]
        for i in range(len(data_2))}

q = {}
for (i, j, k), v in q_0.items():
    key = (trans_dict[i[0]], j, k)
    if key in q:
        q[key] += v
    else:
        q[key] = v
    q[key] = round(q[key], 1)

data_3["地块类型"] = data_3["地块类型"].astype(str)
data_4 = {(data_3["地块类型"][i].strip(), data_3["作物编号"][i],
            2 if data_3["种植季次"][i] == "第二季" else 1):
                [data_3["亩产量/斤"][i], data_3["种植成本/(元/亩)"][i], data_3["销售单价/(元/斤)"][i]]
            for i in range(len(data_3))}
# 1.2 亩产， 成本和价格
#r、c、p的键是地块类型、作物编号、季度组成的元组
#r的值是亩产，c的值是成本，p的值是价格
_, r, c, p = gp.multidict(data_4)#调用gurobipy，从一个数据源创建多个字典
pl = {}
pu = {}
for k, v in p.items():
    try:
        str1 = v.split('-')#把价格区间的两端分别复制给pl、pu
        pl[k] = float(str1[0])
        pu[k] = float(str1[1])
    except:
        pass
    #pl是price_low;pu是price_up
# 1.3 地块⾯积
a = {data_1["地块名称"][i]: data_1["地块面积/亩"][i]
    for i in range(len(data_1))}

# 2. 定义集合
# 2.1 作物
LDR = ["黄豆", "黑豆", "红豆", "绿豆", "爬豆"] # 粮⻝（⾖类）
LD = list(range(1, 6))
LR = ["小麦", "玉米", "谷子", "高粱", "黍子", "荞麦", "南瓜", "红薯", "莜麦", "大麦"] # 粮⻝
L = list(range(6, 16))
LSR = ["水稻"] # 粮⻝⽔浇地
LS = list(range(16, 17))
SDR = ["豇豆", "刀豆", "芸豆"] # 蔬菜（⾖类）
SD = list(range(17, 20))
S_1R = ["土豆", "西红柿", "茄子", "菠菜", "青椒", "菜花", "包菜", "油麦菜", "小青菜","黄瓜", "生菜", "辣椒", "空心菜", "黄心菜", "芹菜"]
S_1 = list(range(20, 35))
S_2R = ["大白菜", "白萝卜", "红萝卜"]
S_2 = list(range(35, 38))
SJR = ["榆黄菇", "香菇", "白灵菇", "羊肚菌"]
SJ = list(range(38, 42))
J = LD + L + LS + SD + S_1 + S_2 + SJ#作物编号集合
J1 = LDR + LR + LSR + SDR + S_1R + S_2R + SJR#作物集合
J_d = dict(zip(J1, J))#J1作为键，J作为值
# 2.2 地块
A = ["A" + str(i) for i in range(1, 7)] # 得到形如“A1，A2...A6”的列表
B = ["B" + str(i) for i in range(1, 15)] # 梯⽥
C = ["C" + str(i) for i in range(1, 7)] # ⼭坡地
D = ["D" + str(i) for i in range(1, 9)] # ⽔浇地
E = ["E" + str(i) for i in range(1, 17)] # 普通⼤棚
F = ["F" + str(i) for i in range(1, 5)] # 智慧⼤棚
trans_dict_2 = {"平旱地": A, "梯田": B, "山坡地": C, "水浇地": D, "普通大棚": E,"智慧大棚": F}
I = A + B + C + D + E + F#地块集合
# 2.3 年份
Y = list(range(2024, 2031))
# 2.4 季节
KR = ["第一季", "第二季"]
K = [1, 2]
# 2.5 集合组合
#按种植耕地分
#product计算笛卡尔积,V是四维变量
V_1 = list(product(A + B + C, LD + L, [K[0]], Y))
V_2 = list(product(D, LS, [K[0]], Y))
V_3 = list(product(D + E + F, SD + S_1, [K[0]], Y))
V_4 = list(product(F, SD + S_1, [K[1]], Y))#智慧大棚的第二季单独出来
V_5 = list(product(D, S_2, [K[1]], Y))
V_6 = list(product(E, SJ, [K[1]], Y))
V = V_1 + V_2 + V_3 + V_4 + V_5 + V_6
#按季节来分
#W是三维变量，表示地块季节集合，与作物无关，用于面积约束
W1 = product(A + B + C + D + E + F, [K[0]], Y)
W2 = product(D + E + F, [K[1]], Y)
W = list(W1) + list(W2)
#地块类型集合
II = list(trans_dict.values())

# 3. 处理参数

# 3.1 销量处理
sale_data = {}
#q是面积，r是亩产，乘起来就是销量
sale_data_2023 = {i: q[i] * r[i]
                    for i in q}#i遍历q的每个键，i是地块、作物编号、季度的元组
#sale_data_2023的键是地块、作物编号、季度的元组，值是销量
# 3.2 亩产处理
mu_data = {}
mu_data_2023 = {i: r[i]
                for i in r}
# 3.3 成本处理
ch_data = {}
ch_data_2023 = {i: c[i]
                for i in c}
# 3.4 价格处理
jia_data_l = {}
jia_data_2023_l = {i: pl[i]
                for i in pl}
jia_data_u = {}
jia_data_2023_u = {i: pu[i]
                for i in pl}
for idx, t in enumerate(Y):#t遍历年份，idx是元素的索引，从0开始
    for i, j, k in q:
        sale_data[i, j, k, t] = sale_data_2023[i, j, k]
    for i, j, k in q:
        sale_data[i, j, k, t] = sale_data_2023[i, j, k]
    for i, j, k in r:
        mu_data[i, j, k, t] = mu_data_2023[i, j, k]
    for i, j, k in c:
        ch_data[i, j, k, t] = ch_data_2023[i, j, k]
    for i, j, k in pl:
        jia_data_l[i, j, k, t] = jia_data_2023_l[i, j, k]
        jia_data_u[i, j, k, t] = jia_data_2023_u[i, j, k]

# 4. 建⽴模型
m = gp.Model("cumcm_2024")
# 4.1 决策变量
x = m.addVars(V, vtype=gp.GRB.CONTINUOUS, lb=0, name="x")
y = m.addVars(V, vtype=gp.GRB.BINARY, lb=0, name="y")
x_s = m.addVars(pl, Y, vtype=gp.GRB.CONTINUOUS, lb=0, name="xs")
l_s = m.addVars(pl, Y, vtype=gp.GRB.BINARY, name="ls")
# 4.2 约束条件
# 1. ⾯积约束

m.addConstrs((x[i, j, k, t] <= a[i] * y[i, j, k, t]
for i, j, k, t in V), "C11")
m.addConstrs((x.sum(i, "*", k, t) <= a[i]
for (i, k, t) in W), "C12")
 # 2. ⽔稻每年⼀季节或者两季节蔬菜

m.addConstrs((y[i, LS[0], K[0], t] + y[i, j, K[0], t] <= 1
 for i in D
 for j in S_1 + SD
 for t in Y), "C21")


m.addConstrs((y[i, LS[0], K[0], t] + y[i, j, K[1], t] <= 1
for i in D
for j in S_2
for t in Y), "C22")
# 3. 重种约束
m.addConstrs((y[i, j, K[0], Y[t]] + y[i, j, K[0], Y[t + 1]] <= 1
for i in A + B + C
for j in LD + L
for t in range(len(Y) - 1)), "C31")

m.addConstrs((y[i, j, K[0], Y[t]] + y[i, j, K[0], Y[t + 1]] <= 1
for i in D
for j in LS
for t in range(len(Y) - 1)), "C32")

m.addConstrs((y[i, j, K[0], Y[t]] + y[i, j, K[1], Y[t]] <= 1
for i in F
for j in SD + S_1
for t in range(len(Y))), "C33")

m.addConstrs((y[i, j, K[1], Y[t]] + y[i, j, K[0], Y[t + 1]] <= 1
for i in F
for j in SD + S_1
for t in range(len(Y) - 1)), "C34")

# 4. 所有⼟地三年内⾄少种植⼀次⾖类作物
m.addConstrs((gp.quicksum(y[i, j, K[0], Y[tau]]
for j in LD
for tau in range(t, t + 3)) >= 1
for i in A + B + C
for t in range(0, len(Y) - 2)), "C41")

m.addConstrs((gp.quicksum(y[i, j, K[0], Y[tau]]
for j in SD
for tau in range(t, t + 3)) >= 1
for i in D + E
for t in range(0, len(Y) - 2)), "C42")

m.addConstrs((gp.quicksum(y[i, j, k, Y[tau]]
for k in K
for j in SD
for tau in range(t, t + 3)) >= 1
for i in F
for t in range(0, len(Y) - 2)), "C43")

# 5. 种植地不能太分散
theta = 5
for idx, I_ in enumerate(list(trans_dict_2.values())):
    m.addConstrs((gp.quicksum(y.get((i, j, k, t), 0)
                            for i in I_) <= theta
                for j in J
                for k in K
                for t in Y), "C5"+str(idx))
    
# 6. 单个地块 （含⼤棚 种植 的 ⾯积不宜太⼩，等等

delta = 1 / 3
m.addConstrs((x[i, j, k, t] >= a[i] * y[i, j, k, t] * delta
                for i, j, k, t in V), "C6")

m.addConstrs((x_s[I0, j, k, t] == l_s[I0, j, k, t] * x.sum(trans_dict_2[I0], j, k, t) * mu_data[I0, j, k, t] +(1 - l_s[I0, j, k, t]) * sale_data.get((I0, j, k, t), 1e10)
                for I0, j, k, t in x_s), "C71")
m.addConstrs(((l_s[I0, j, k, t] == 1) >> (x.sum(trans_dict_2[I0], j, k, t) * mu_data[I0, j, k, t] <=sale_data.get((I0, j, k, t), 1e10))
                for I0, j, k, t in x_s), "C72")

m.addConstrs(((l_s[I0, j, k, t] == 0) >> (x.sum(trans_dict_2[I0], j, k, t) * mu_data[I0, j, k, t] >=sale_data.get((I0, j, k, t), 1e10))
                for I0, j, k, t in x_s), "C73")

# 4.3 ⽬标函数
obj = gp.LinExpr()
for t in Y:
    for I0, j, k in pl:
        if I0 == "平旱地":
            p_c = jia_data_l[I0, j, k, t] * 0.75 + 0.25 * jia_data_u[I0, j, k,t]
        if I0 == "梯⽥":
            p_c = jia_data_l[I0, j, k, t] * 0.75 + 0.25 * jia_data_u[I0, j, k,t]
        if I0 == "⼭坡地":
            p_c = jia_data_l[I0, j, k, t] * 0.75 + 0.25 * jia_data_u[I0, j, k,t]
        if I0 == "⽔浇地":
            p_c = jia_data_l[I0, j, k, t] * 0.75 + 0.25 * jia_data_u[I0, j, k,t]
        if I0 == "普通⼤棚":
            p_c = jia_data_l[I0, j, k, t] * 0.75 + 0.25 * jia_data_u[I0, j, k,t]
        if I0 == "智慧⼤棚":
            p_c = jia_data_l[I0, j, k, t] * 0.75 + 0.25 * jia_data_u[I0, j, k,t]
                
        c_c = ch_data[I0, j, k, t]
        current = p_c * x_s[I0, j, k, t] - c_c * x.sum(trans_dict_2[I0], j, k,t)
        obj += current

m.setObjective(obj, gp.GRB.MAXIMIZE)

# 5. 模型求解
m.write("model.lp")
m.setParam(gp.GRB.Param.MIPGap, 0.01)
m.setParam(gp.GRB.Param.TimeLimit, 120)
m.optimize()

# 6. 结果输出
xv = m.getAttr("X", x)
yv = m.getAttr("X", y)
# for k, v in xv.items():
# if v < .001 or k[3] > 2024:
# continue
# print(k, ":", v)

workbook = load_workbook(r"C:\Users\86181\Desktop\数模作业、笔记\农作物种植策略\results\result1_1.xlsx")
for t in Y:
    data_r = pd.read_excel(r"C:\Users\86181\Desktop\数模作业、笔记\农作物种植策略\results\result1_1.xlsx", sheet_name=str(t))
    # writer = xlsxwriter.Workbook('existing_file.xlsx', engine='openpyxl')
    worksheet_writer = workbook[str(t)]
    for index, row in data_r.iterrows():
        r_data = []
        for cindex, col_name in enumerate(data_r.columns[2:]):
            cc_name = col_name.strip()
            if (data_r["地块名"][index], J_d[cc_name], K[0] if index <= 53 else K[1], t) in xv:
                 vv = xv[data_r["地块名"][index], J_d[cc_name], K[0] if index <= 53 else K[1], t]
                 if vv > 0.001:
                     worksheet_writer.cell(index + 2, 3 + cindex, vv)

                     file_path =r"C:\Users\86181\Desktop\数模作业、笔记\农作物种植策略\results\result1_2.xlsx"
                     if os.path.exists(file_path):
                         os.remove(file_path)
                     workbook.save(file_path)